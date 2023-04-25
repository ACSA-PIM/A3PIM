from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill  # 导入填充模块
import global_variable as glv
import time
from multiBar import time2String
from data import dataDictInit
from tsjPython.tsjCommonFunc import *
import string

# https://blog.csdn.net/David_Dai_1108/article/details/78702032
def readExcelSet(pageName, columnName):
    yellowPrint("Reading excel {} data……".format(columnName))
    unique_revBiblock=set()
    if not columnName=="block_binary":
        return unique_revBiblock
    from openpyxl import load_workbook
    # 只读模式打开文件
    wb = load_workbook(glv._get("HistoryDataFile"), read_only=True)
    # 获得所有 sheet 的名称()
    name_list = wb.sheetnames
    # 根据 sheet 名字获得 sheet
    if pageName not in name_list:
        ic(pageName,"Not existed")
        return unique_revBiblock
    ic(pageName,"existed")
    my_sheet = wb[pageName]
    for rowNum in range(2,my_sheet.max_row):
        ic(rowNum)
        unique_revBiblock.add(my_sheet["B"+str(rowNum)].value)
    return unique_revBiblock

def colorAccuracyEntry(columnLetter, accuracyEntry, tmp_block_binary_reverse ,lineNum, ws):
    toFill=ws['{}{}'.format(columnLetter, lineNum+1)]
    if accuracyEntry[tmp_block_binary_reverse] > 1:
        toFill.fill=PatternFill('solid', fgColor='FF0000') # 红
    elif accuracyEntry[tmp_block_binary_reverse] > 0.5:
        toFill.fill=PatternFill('solid', fgColor='ffc7ce') # 浅红
    elif accuracyEntry[tmp_block_binary_reverse] > 0.25:
        toFill.fill=PatternFill('solid', fgColor='FFFF00') # 黄
    elif accuracyEntry[tmp_block_binary_reverse] > 0.1:
        toFill.fill=PatternFill('solid', fgColor='FFFF99') # 浅黄

def excelGraphInit():
    wb = Workbook()
    ws = wb.active # 找当前Sheet
    ws.title = 'Graph'
    ws = wb["Graph"]
    for i in string.ascii_uppercase[:14]:
        ws.column_dimensions[i].width = 15 # 修改列宽
    ws.append(["Applications","CPU-ONLY","PIM-ONLY","Instruction", 'MPKI-based',\
                "Architecture-Suitability/Greedy","PIMProf"])
    return wb

def excelGraphAdd(wb,writeList):
    ws = wb["Graph"]
    ws.append(writeList)
    wb.save(glv._get("excelOutPath"))
    
def excelGraphBuild(wb,processBeginTime):
    # 一个图两个轴
    ws = wb["Graph"]
    ct_bar = BarChart()
    taskNum=len(glv._get("taskList"))
    ws['D1'] = '误差比值'
    for i in range(2, taskNum+2):
        toFill=ws[f'B{i}']
        toFillValue=ws[f'B{i}'].value
        if toFillValue > 0.5:
            toFill.fill=PatternFill('solid', fgColor='FF0000') #红
        elif toFillValue > 0.2:
            toFill.fill=PatternFill('solid', fgColor='ffc7ce') #浅红
        elif toFillValue > 0.1:
            toFill.fill=PatternFill('solid', fgColor='FFFF00') #黄
        elif toFillValue <= 0.1:
            toFill.fill=PatternFill('solid', fgColor='c6efce') #绿
        toFill=ws[f'C{i}']
        toFillValue=ws[f'C{i}'].value
        if toFillValue > 0.5:
            toFill.fill=PatternFill('solid', fgColor='FF0000') #红
        elif toFillValue > 0.2:
            toFill.fill=PatternFill('solid', fgColor='ffc7ce') #浅红
        elif toFillValue > 0.1:
            toFill.fill=PatternFill('solid', fgColor='FFFF00') #黄
        elif toFillValue <= 0.1:
            toFill.fill=PatternFill('solid', fgColor='c6efce') #绿


        if ws[f'B{i}'].value==0:
            ws[f'D{i}']=0
        else:
            ratio = ws[f'C{i}'].value / ws[f'B{i}'].value
            ws[f'D{i}'] = ratio
            toFill=ws[f'D{i}']
            if ratio > 2:
                toFill.fill=PatternFill('solid', fgColor='c6efce') #绿
            elif ratio < 1:
                toFill.fill=PatternFill('solid', fgColor='ffc7ce') #红

            # https://openpyxl.readthedocs.io/en/stable/styles.html
    d_ref = Reference(ws, min_col=2, min_row=1, max_row=taskNum+1, max_col=3)
    ct_bar.add_data(d_ref, titles_from_data=True)
    series = Reference(ws, min_col=1, min_row=2, max_row=taskNum+1)
    ct_bar.set_categories(series)
    ct_bar.x_axis.title = '应用'
    ct_bar.y_axis.title = '误差'
    ct_bar.y_axis.majorGridlines = None
    ct_bar.title = '各应用静态分析误差对比表'
    ct_line = LineChart()
    d_ref = Reference(ws, min_col=4, min_row=1, max_row=taskNum+1)
    ct_line.add_data(d_ref, titles_from_data=True)
    ct_line.y_axis.axId = 200 # 不为空即可
    ct_line.y_axis.title = '误差比值'
    # 让线条和第一图的最大值相交
    ct_line.y_axis.crosses = 'max'
    ct_bar += ct_line # 只支持+=赋值，不能直接+
    ws.add_chart(ct_bar, 'A30')
    ws.append(["time spent {}".format(time2String(int(time.time()-processBeginTime)))])
    excelAddHeatmap(ws)
    wb.save(glv._get("excelOutPath"))

def excelAddHeatmap(ws):
    position=1
    taskList = glv._get("taskList")
    for taskKey, taskName in taskList.items():
        img = Image("./pictures/"+taskName+".png")
        img.width, img.height=(300,3*90)
        ws.add_image(img, 'L'+str(position))
        img = Image("./pictures/"+taskName+"_baseline.png")
        img.width, img.height=(300,3*90)
        ws.add_image(img, 'P'+str(position))
        position+=15